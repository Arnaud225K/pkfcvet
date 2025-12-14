# -*- coding: utf-8 -*-
from django.shortcuts import render, get_object_or_404
from django.views.generic.base import TemplateView
from django.core.mail import send_mail
import xlrd
import datetime
import os
import re
import openpyxl
from transliterate import translit
from django.http import Http404
from django.contrib.redirects.models import Redirect
from menu.models import MenuCatalog, Product, tableProductDop, ProductUsluga
from menu.views import import_image, set_param_category
from Marochnik.models import MarkaMetall
from GOST.models import GOST

from pkfcvet.settings import EMAIL_HOST_USER, HOST_NAME, SITE_NAME
from project_settings.models import ProjectSettings
from .models import ImportData, ExportData
from .forms import UploadFileForm



DEFAULT_OPTION = u'Все'

IND_STATE_WORK = 1
IND_STATE_READY = 2
IND_STATE_ERROR = 3
IND_STATE_POTENTIAL_ERROR = 4

ID_TYPE_MENU_CATALOG = 1

COUNT_ROW_FOR_INFO_UPDATE = 100
COUNT_ROW_FOR_INFO_UPDATE_CONTROL = 100
# COUNT_ROW_FOR_INFO_UPDATE_FORMAT = 1000

MAX_COUNT_ROW_INFO = 1000

TYPE_ACTION_IMPORT = "import"
TYPE_ACTION_IMPORT_MARKA = "import_marka"
TYPE_ACTION_IMPORT_HARD = "import_hard"
TYPE_ACTION_CONTROL = "control"


DEFAULT_MESSAGE_ERROR = u"Импорт завершен с ошибками"

COLUMN_ID = 0
COLUMN_ORDER = 1
COLUMN_NAME = 2
COLUMN_SIZE_A = 3
COLUMN_SIZE_B = 4
COLUMN_SIZE_C = 5
COLUMN_SIZE_D = 6
COLUMN_SIZE_E = 7
COLUMN_SIZE_F = 8
COLUMN_SIZE_L = 9
COLUMN_MARKA = 10
COLUMN_GOST = 11
COLUMN_AVAILABLE = 12
COLUMN_PRICE = 13
COLUMN_ED_IZM = 14
COLUMN_IMAGE_1 = 15
COLUMN_IMAGE_2 = 16
COLUMN_DESCRIPTION = 17
COLUMN_VENDOR = 18
COLUMN_COUNTRY_VENDOR = 19
COLUMN_METHOD = 20
COLUMN_CATEGORY_ID = 21
COLUMN_SERTIFICATE_ID = 22
COLUMN_PRODUCT_1_ID = 23
COLUMN_PRODUCT_2_ID = 24
COLUMN_PRODUCT_3_ID = 25
COLUMN_PRODUCT_4_ID = 26
COLUMN_USLUGA_1_ID = 27
COLUMN_USLUGA_2_ID = 28
COLUMN_USLUGA_3_ID = 29
COLUMN_USLUGA_4_ID = 30

COLUMN_IMAGE_1_HTTP = -2
COLUMN_IMAGE_2_HTTP = -1
COLUMN_IMAGE_1_TECH_NAME = -4
COLUMN_IMAGE_2_TECH_NAME = -3

COLUMN_COUNT = 31
ROW_MIN_COUNT = 2

CONTROL_CODE = "73aoF6N"

FLAG_IMPORT = 1


def static_admin_url(request):
    return {
        'static_admin_url': '/static/admin_m/',
    }


class AdminMIndexView(TemplateView):
    """
    Класс для отображения главной страницы
    """
    template_name = "admin_m/index.html"

    def get(self, request):
        return render(request, self.template_name, locals())


class AdminMImportInfoView(TemplateView):
    """
    Класс для отображения страницы информации об импорте
    """
    template_name = "admin_m/control/import/import_info.html"

    def get(self, request, import_info_slug):
        current_info = get_object_or_404(ImportData, id=import_info_slug)
        current_info_information = current_info.info.replace("\n", "<br/>")
        return render(request, self.template_name, locals())


class AdminMImportView(TemplateView):
    """
    Класс для отображения страницы импорта
    """
    template_name = "admin_m/control/import/import.html"

    def get(self, request):
        items_list = ImportData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO

        return render(request, self.template_name, locals())

    def post(self, request, template_name="admin_m/control/import/import_table.html"):
        post_data = request.POST
        email = None
        try:
            email = post_data["email"]
        except KeyError:
            pass

        form = UploadFileForm(request.POST, request.FILES)
        file_name = None
        if form.is_valid():
            main_name = translit(request.FILES['file'].name.strip(), "ru", reversed=True)
            file_name = 'import/files/import_' + get_today() + '_' + main_name
            handle_uploaded_file(request.FILES['file'], file_name)
        else:
            raise Http404()

        try:
            type_action = post_data["type_action"]
        except KeyError:
            raise Http404()

        flag_hard_import = None
        try:
            flag_hard_import = post_data["flag_hard_import"]
        except KeyError:
            pass

        action_str = ""
        if type_action == TYPE_ACTION_IMPORT:
            if flag_hard_import:
                action_str = u"Импорт (принудительный)"
                type_action = TYPE_ACTION_IMPORT_HARD
            else:
                action_str = u"Импорт"
        elif type_action == TYPE_ACTION_CONTROL:
            action_str = u"Проверка дубликатов"
        elif type_action == TYPE_ACTION_IMPORT_MARKA:
            action_str = u"Импорт марок"

        info = ImportData()
        info.name = request.FILES['file'].name
        info.action = action_str
        info.user = request.user.get_full_name()
        info.email = email
        info.file = file_name
        info.state_id = IND_STATE_WORK
        info.result = "0 %"
        info.result_percent = 0
        info.info = u""
        info.save()

        import threading
        t = threading.Thread(target=import_data, args=[info, request.scheme, type_action])
        t.setDaemon(True)
        t.start()
        # import_data(info, type_action)


        items_list = ImportData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO

        return render(request, template_name, locals())


def send_mail_result(email, user, name, link, link_admin, action, result):
    subject = u"Панель управления %s. %s (файл: %s) - %s" % (SITE_NAME, action, name, result)
    plain_text = u"""Здравствуйте, %s!\n%s (файл: %s) %s.""" % (user, action, name, result)
    html_text = u"""<h1>Здравствуйте, %s!</h1>\n<p>%s файла: %s - %s. </p>\n\n
        <p>Вы можете посмотреть подробнее по <a href='%s'>прямой ссылке</a> или из <a href='%s'>панели управления</a></p>""" % (user, action, name, result, link, link_admin)

    send_mail(
        subject,
        plain_text,
        EMAIL_HOST_USER,
        [email],
        html_message=html_text,
        fail_silently=False,
    )


def get_today():
    return datetime.datetime.today().strftime("%Y-%m-%d %H:%M")


def handle_uploaded_file(f, file_name):
    filename = 'www/media/' + file_name
    with open(filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def import_data(info, scheme, type_action):
    data_load = load_xls(info)
    if info.state_id == IND_STATE_ERROR:
        info.result = DEFAULT_MESSAGE_ERROR
        info.info += DEFAULT_MESSAGE_ERROR + "\n"
        info.save()
        return

    data_format, new_marka_list = format_type(info, data_load, type_action)

    if type_action == TYPE_ACTION_IMPORT:
        if info.state_id == IND_STATE_WORK:
            import_data_import(info, data_format)
    elif type_action == TYPE_ACTION_IMPORT_HARD:
        if info.state_id == IND_STATE_WORK:
            import_data_import_hard(info, data_format)
    elif type_action == TYPE_ACTION_CONTROL:
        import_data_control(info, data_format)
    elif type_action == TYPE_ACTION_IMPORT_MARKA:
        import_data_import_marka(info, new_marka_list)

    if info.state_id == IND_STATE_ERROR:
        info.result = "Завершен с ошибками"
        info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, устраните ошибки и повторите импорт"
        info.save()
    elif info.state_id == IND_STATE_POTENTIAL_ERROR:
        info.result = "Завершен с ошибками"
        info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, если Вы действительно хотите изменить категорию у существующих позиций, установите флаг 'отключить проверку дубликатов' и повторите импорт"
        info.save()
    else:
        info.state_id = IND_STATE_READY
        info.save()

    if info.email:
        url_site = scheme + "://" + HOST_NAME
        link = url_site + info.get_absolute_url()
        link_admin = url_site + "/admin_m/import/"
        send_mail_result(info.email, info.user, info.name, link, link_admin, info.action, info.state.name)
        info.info += u""
        info.save()


def load_xls(info):
    file_name = info.file.path
    book = xlrd.open_workbook(file_name)
    s = book.sheet_by_index(0)
    data_load = []
    index_row = 0
    if s.nrows < ROW_MIN_COUNT:
        info.info += u"Неверный формат файла: количество строк меньше минимального - %d\n" % ROW_MIN_COUNT
        info.state_id = IND_STATE_ERROR
    if s.ncols != COLUMN_COUNT:
        info.info += u"Неверный формат файла: количество столбцов в файле %d, а должно быть %d\n" % (s.ncols, COLUMN_COUNT)
        info.state_id = IND_STATE_ERROR
    for row in range(s.nrows):
        values = []
        for col in range(s.ncols):
            values.append(s.cell(row, col).value)
        data_load.append(values)
    return data_load


def get_marka_name(marka):
    key = marka
    return key


def get_gost_name(gost, dict_gost):
    gost_5 = ""
    if not gost == "":
        key = gost.replace(' ', '')
        try:
            gost_5 = dict_gost[key]
        except:
            try:
                gost_5 = dict_gost[u"ГОСТ" + key]
            except:
                try:
                    gost_5 = dict_gost[u"ГОСТР" + key]
                except:
                    try:
                        gost_5 = dict_gost[u"ГОСТ" + key + '-78']
                    except:
                        try:
                            gost_5 = dict_gost[u"ГОСТ" + key + '-87']
                        except:
                            try:
                                gost_5 = dict_gost[u"ТУ" + key]
                            except:
                                return False
    return gost_5


def control_float_type(info, i, test_size, size_name):
    if test_size:
        try:
            float(test_size)
        except ValueError:
            info.info += u"%d строка: %s должен быть числом: %s\n" % (i, size_name, test_size)
            info.state_id = IND_STATE_ERROR
            info.save()


def format_type(info, data_load, type_action):
    data_format = []
    new_marka_list = []
    new_gost_list = []
    rez = True

    category_list = []
    category_list_map = MenuCatalog.objects.all().only('id').values('id')
    for item in category_list_map:
        category_list.append(item['id'])
    marka_dict = {}
    marka_dict_map = MarkaMetall.objects.all().only('id', 'name').values('id', 'name')
    for item in marka_dict_map:
        marka_dict[item['name']] = item['id']

    gost_dict = {}
    gost_dict_map = GOST.objects.all().only('id', 'number').values('id', 'number')
    for item in gost_dict_map:
        gost_dict[item['number'].replace(' ', '')] = item['id']

    info.result = "Подготовка"
    # info.result = "Подготовка - 0 %"
    # count_items = len(data_load) - 1
    for i in range(1, len(data_load)):
        # info.result_percent += (1. / count_items * 100)
        # print("format %.2f  %d / %d" % (info.result_percent, i, count_items))
        # if i % COUNT_ROW_FOR_INFO_UPDATE_FORMAT == 0 or i == count_items:
        #     info.result = ("Подготовка - %.2f" % info.result_percent) + " %"
        #     info.save()

        data_format.append([])
        data_format[i - 1].append(str(data_load[i][COLUMN_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_ORDER])[:-2])
        data_format[i - 1].append(data_load[i][COLUMN_NAME])
        
        array_category = str(data_load[i][COLUMN_CATEGORY_ID]).split(";")
        for a_category in array_category:
            category_id = ""
            tmp_category_id = ""
            category_id = a_category.replace(".0", "")

            try:
                tmp_category_id = int(category_id)
                category_id = tmp_category_id
            except:
                pass

            if not (category_id in category_list):
                info.info += u"%d строка: Отсутствует категория с идентификатором: %s\n" % (i, category_id)
                info.state_id = IND_STATE_ERROR
                info.save()

        # control_float_type(info, i, data_load[i][COLUMN_SIZE_A], u'Размер А')
        # control_float_type(info, i, data_load[i][COLUMN_SIZE_B], u'Размер B')
        # control_float_type(info, i, data_load[i][COLUMN_SIZE_C], u'Размер C')
        # control_float_type(info, i, data_load[i][COLUMN_SIZE_D], u'Размер D')
        # control_float_type(info, i, data_load[i][COLUMN_SIZE_E], u'Размер E')
        # control_float_type(info, i, data_load[i][COLUMN_SIZE_F], u'Размер F')


        data_format[i - 1].append(data_load[i][COLUMN_CATEGORY_ID])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_A])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_B])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_C])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_D])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_E])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_F])
        data_format[i - 1].append(data_load[i][COLUMN_SIZE_L])
        marka_name = get_marka_name(data_load[i][COLUMN_MARKA])
        if marka_name:
            marka_id = ''
            try:
                marka_id = marka_dict[marka_name]
                data_format[i - 1].append(marka_id)
            except:
                # try:
                info.info += u"%d строка: Отсутствует марка: %s\n" % (i, marka_name)
                # except:
                #     info.info += str(i) + " строка: Отсутствует марка содержащая спецсимволы\n"
                info.state_id = IND_STATE_ERROR
                info.save()
                if not (marka_name in new_marka_list):
                    new_marka_list.append(marka_name)
                data_format[i - 1].append('')
        else:
            data_format[i - 1].append(marka_name)

        if data_load[i][COLUMN_GOST]:
            gost_id = get_gost_name(data_load[i][COLUMN_GOST], gost_dict)
            if gost_id:
                data_format[i - 1].append(gost_id)
            else:
                info.info += u"%d строка: Отсутствует ГОСТ, ТУ: %s\n" % (i, data_load[i][COLUMN_GOST])
                info.state_id = IND_STATE_ERROR
                info.save()
                if not (data_load[i][COLUMN_GOST] in new_gost_list):
                    new_gost_list.append(data_load[i][COLUMN_GOST])
        else:
            data_format[i - 1].append('')

        data_format[i - 1].append(data_load[i][COLUMN_AVAILABLE])
        data_format[i - 1].append(data_load[i][COLUMN_PRICE])
        if not str(data_load[i][COLUMN_PRICE]):
            info.info += u"%d строка: Отсутствует цена!\n" % (i)
            info.state_id = IND_STATE_ERROR
            info.save()

        data_format[i - 1].append(data_load[i][COLUMN_ED_IZM])

        image_name_1 = ""
        image_name_1_http = data_load[i][COLUMN_IMAGE_1]
        image_name_2 = ""
        image_name_2_http = data_load[i][COLUMN_IMAGE_2]

        if data_load[i][COLUMN_IMAGE_1]:
            # rx = re.compile('[\/]+')
            rx = re.compile(r'[\/]+')
            rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_1])
            if rezImg[-1]:
                image_name_1 = rezImg[-1]
            elif rezImg[-2]:
                image_name_1 = rezImg[-2]

        if data_load[i][COLUMN_IMAGE_2]:
            # rx = re.compile('[\/]+')
            rx = re.compile(r'[\/]+')
            rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_2])
            if rezImg[-1]:
                image_name_2 = rezImg[-1]
            elif rezImg[-2]:
                image_name_2 = rezImg[-2]

        image_name_bd = "uploads/images/load/" + image_name_1
        data_format[i - 1].append(image_name_bd)
        image_name_bd = "uploads/images/load/" + image_name_2
        data_format[i - 1].append(image_name_bd)

        data_format[i - 1].append(data_load[i][COLUMN_DESCRIPTION])
        data_format[i - 1].append(data_load[i][COLUMN_VENDOR])
        data_format[i - 1].append(data_load[i][COLUMN_COUNTRY_VENDOR])
        data_format[i - 1].append(data_load[i][COLUMN_METHOD])

        data_format[i - 1].append(str(data_load[i][COLUMN_SERTIFICATE_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_PRODUCT_1_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_PRODUCT_2_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_PRODUCT_3_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_PRODUCT_4_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_USLUGA_1_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_USLUGA_2_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_USLUGA_3_ID])[:-2])
        data_format[i - 1].append(str(data_load[i][COLUMN_USLUGA_4_ID])[:-2])

        data_format[i - 1].append(image_name_1)
        data_format[i - 1].append(image_name_2)
        data_format[i - 1].append(image_name_1_http)
        data_format[i - 1].append(image_name_2_http)

    info.result_percent = 0
    info.save()
    if new_marka_list:
        info.info += u"\nОтсутствуют марки:\n"
        for item in new_marka_list:
            info.info += u"%s\n" % item
        info.save()
        info.info += u"Вы можете загрузить их через 'Импорт марок'\n"
    if new_gost_list:
        info.info += u"\nОтсутствуют ГОСТы:\n"
        for item in new_gost_list:
            info.info += u"%s\n" % item
        info.save()
    return data_format, new_marka_list



def import_data_import_hard(info, data_format):
    list_import_image = []
    data_format_category = []
    count_products = len(data_format)
    for i in range(len(data_format)):
        product = Product()
        info.result_percent += (1. / count_products * 100)
        # print("import hard %.2f   %d/%d" % (info.result_percent, i, count_products))
        if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
            info.result = ("%.2f" % info.result_percent) + " %"
            info.save()

        product.id = data_format[i][COLUMN_ID]
        if data_format[i][1]:
            product.order_number = data_format[i][1]
        product.name_main = data_format[i][2]

        product.catalogOne_id = None
        product.catalogTwo_id = None
        product.catalogThree_id = None

        array_category = str(data_format[i][3]).split(";")
        for a_category in array_category:
            iii_cat = a_category.replace(".0", "")
            if len(array_category) == 1:
                if array_category[0] == a_category:
                    product.catalog_id = int(iii_cat)
            if len(array_category) == 2:
                if array_category[0] == a_category:
                    product.catalog_id = int(iii_cat)
                if array_category[1] == a_category:
                    product.catalogOne_id = int(iii_cat)
            if len(array_category) == 3:
                if array_category[0] == a_category:
                    product.catalog_id = int(iii_cat)
                if array_category[1] == a_category:
                    product.catalogOne_id = int(iii_cat)
                if array_category[2] == a_category:
                    product.catalogTwo_id = int(iii_cat)
            if len(array_category) == 4:
                if array_category[0] == a_category:
                    product.catalog_id = int(iii_cat)
                if array_category[1] == a_category:
                    product.catalogOne_id = int(iii_cat)
                if array_category[2] == a_category:
                    product.catalogTwo_id = int(iii_cat)
                if array_category[3] == a_category:
                    product.catalogThree_id = int(iii_cat)

            if not (int(iii_cat) in data_format_category):
                data_format_category.append(int(iii_cat))

        if data_format[i][4]:
            product.size_a = data_format[i][4]
        if data_format[i][5]:
            product.size_b = data_format[i][5]
        if data_format[i][6]:
            product.size_c = data_format[i][6]
        if data_format[i][7]:
            product.size_d = data_format[i][7]
        if data_format[i][8]:
            product.size_e = data_format[i][8]
        if data_format[i][9]:
            product.size_f = data_format[i][9]
        if data_format[i][10]:
            product.size_l = data_format[i][10]
        if data_format[i][11]:
            product.marka_id = data_format[i][11]
        if data_format[i][12]:
            product.gost_id = data_format[i][12]
        product.available = data_format[i][13]
        product.price = data_format[i][14]
        product.ed_izm = data_format[i][15]
        product.image = data_format[i][16]
        product.image_dop = data_format[i][17]
        product.description = data_format[i][18]
        product.vendor = data_format[i][19]
        product.vendor_country = data_format[i][20]
        product.method_of_manufacture = data_format[i][21]
        product.isHidden = False
        if data_format[i][22]:
            product.sertificate_id = data_format[i][22]

        if product.slug is None or product.slug == "":
            product.set_slug()

        # Контроль на дубли позиций по латинским именам
        try:
            product_tmp = Product.objects.only('id').get(slug=product.slug)
            if product_tmp.id != int(product.id):
                info.info += u"ОШИБКА при сохранении продукта - дубликат латинского имени '%s' (текущий id=%d, существующий id=%d)\n" % (product.slug, int(product.id), product_tmp.id)
                info.state_id = IND_STATE_ERROR
                info.save()
                break
        except:
            pass

        try:
            product.save()
        except Exception as e:
            info.info += u"ОШИБКА при сохранении продукта id=%d: %s\n" % (int(product.id), e)
            info.state_id = IND_STATE_ERROR
            info.save()
            break

        for tmp_i in range(23, 27):
            if data_format[i][tmp_i]:
                product_tableProductDop = tableProductDop()
                product_tableProductDop.product = product.id
                product_tableProductDop.product_dop = data_format[i][tmp_i]
                product_tableProductDop.save()
        for tmp_i in range(27, 31):
            if data_format[i][tmp_i]:
                product_ProductUsluga = ProductUsluga()
                product_ProductUsluga.product = product.id
                product_ProductUsluga.usluga = data_format[i][tmp_i]
                product_ProductUsluga.save()

        if data_format[i][COLUMN_IMAGE_1_TECH_NAME]:
            tmp_import_image = data_format[i][COLUMN_IMAGE_1_TECH_NAME]
            if 'uploads/images/load/' not in data_format[i][COLUMN_IMAGE_1_HTTP]:
                if not (tmp_import_image in list_import_image):
                    # print("import_image %s %s %s" % (data_format[i][COLUMN_IMAGE_1_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_1_HTTP]))
                    import_image(None, data_format[i][COLUMN_IMAGE_1_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_1_HTTP], info, product.id)
                    list_import_image.append(tmp_import_image)

        if data_format[i][COLUMN_IMAGE_2_TECH_NAME]:
            tmp_import_image = data_format[i][COLUMN_IMAGE_2_TECH_NAME]
            if 'uploads/images/load/' not in data_format[i][COLUMN_IMAGE_2_HTTP]:
                if not (tmp_import_image in list_import_image):
                    # print("import_image %s %s %s" % (data_format[i][COLUMN_IMAGE_2_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_2_HTTP]))
                    import_image(None, data_format[i][COLUMN_IMAGE_2_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_2_HTTP], info, product.id)
                    list_import_image.append(tmp_import_image)

        if info.state_id == IND_STATE_ERROR:
            break

    if info.state_id != IND_STATE_ERROR:
        for item in data_format_category:
            set_param_category(None, item, CONTROL_CODE)

        info.result = "100 %"
        info.save()


def import_data_import(info, data_format):
    info.result = "Проверка id категорий - 0 %"
    count_products = len(data_format)
    for i in range(len(data_format)):
        info.result_percent += (1. / count_products * 100)
        # print("control id %.2f   %d/%d" % (info.result_percent, i, count_products))
        if i % COUNT_ROW_FOR_INFO_UPDATE_CONTROL == 0 or i + 1 == count_products:
            info.result = ("Проверка id категорий - %.2f" % info.result_percent) + " %"
            info.save()

        try:
            product = Product.objects.only('id').get(id=data_format[i][COLUMN_ID])
        except:
            continue
        array_category = str(data_format[i][3]).split(";")
        ii_cat = array_category[0].replace(".0", "")
        if product.catalog_id != int(ii_cat):
            info.info += u"%d строка: Внимание! Возможно ошибочно указан id категории %d (сейчас на сайте у данного продукта id категории %d)\n" % (i, int(ii_cat), product.catalog_id)
            if info.state_id == IND_STATE_WORK:
                info.state_id = IND_STATE_POTENTIAL_ERROR
    info.result_percent = 0
    info.save()
    if info.state_id == IND_STATE_WORK:
        import_data_import_hard(info, data_format)


def import_data_control(info, data_format):
    count_products = len(data_format)
    for i in range(len(data_format)):
        info.result_percent += (1. / count_products * 100)
        if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
            info.result = ("%.2f" % info.result_percent) + " %"
            info.save()

        try:
            product = Product.objects.get(id=data_format[i][COLUMN_ID])
        except:
            continue

        info.info += u"%d строка: Указанный id уже существует (id - %s) \n" % (i, data_format[i][COLUMN_ID])
        info.state_id = IND_STATE_ERROR
        info.save()


def import_data_import_marka(info, new_marka_list):
    for item in new_marka_list:
        marka = MarkaMetall()
        marka.position_id = 22
        marka.name = item
        marka.isHidden = False
        marka.save()

    info.result = "100 %"
    info.state_id = IND_STATE_WORK
    info.info += u"\nМарки успешно загружены"
    info.save()


class AdminMExportView(TemplateView):
    """
    Класс для отображеия страницы импорта
    """
    template_name = "admin_m/control/export/export.html"

    def get(self, request):
        default_option = DEFAULT_OPTION
        category_list = MenuCatalog.objects.filter(typeMenu_id=ID_TYPE_MENU_CATALOG).only('name',).order_by('name')
        items_list = ExportData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO
        return render(request, self.template_name, locals())

    def post(self, request, template_name="admin_m/control/export/export_table.html"):
        post_data = request.POST
        name = None
        email = None
        redirect = None

        try:
            name = post_data["name"]
            email = post_data["email"]
            redirect = post_data["redirect"]
        except KeyError:
            pass

        info = ExportData()
        info.name = name
        info.redirect = redirect
        info.user = request.user.get_full_name()
        info.email = email
        info.state_id = IND_STATE_WORK
        info.result = "0 %"
        info.result_percent = 0
        info.save()

        import threading
        t = threading.Thread(target=export_data, args=[info, request.scheme])
        t.setDaemon(True)
        t.start()

        items_list = ExportData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO

        return render(request, template_name, locals())


def export_all_category(info, url_site):
    global COUNT_ROW_FOR_INFO_UPDATE
    COUNT_ROW_FOR_INFO_UPDATE = 100
    category_list = MenuCatalog.objects.filter(typeMenu_id=ID_TYPE_MENU_CATALOG).only('id', 'name',)
    count_products = Product.objects.all().count()

    d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
    name_all_dir = "all_" + d
    os.system("cd www/media/export/;mkdir {0}".format(name_all_dir))
    # info.info = "<p>Начало экспорта всех категорий</p>"
    for item in category_list:
        # info.info += "<p>СТАРТ - Экспорт категории: %s (id = %d) процент: %.2f</p>" % (item.name, item.id, info.result_percent)
        info.save()
        export_data_category(info,
                             url_site,
                             current_category=item,
                             count_products=count_products,
                             name_all_dir=name_all_dir + "/")
        # info.info += "<p>КОНЕЦ - Экспорт категории: %s (id = %d) процент: %.2f</p>" % (item.name, item.id, info.result_percent)
        info.save()
    info.result = "Архивация"
    info.save()
    os.system("cd www/media/export/;rm {0}.tar.gz;tar -zcvf {0}.tar.gz {0}".format(name_all_dir))
    os.system("rm -R www/media/export/{0}".format(name_all_dir))
    info.link = name_all_dir + ".tar.gz"
    info.result = "100 %"
    info.result_percent = 100
    info.save()


def export_data_category(info, url_site, current_category=None, count_products=None, name_all_dir=""):
	flag_all = True

	if not current_category:
		current_category = MenuCatalog.objects.get(name=info.name)
		flag_all = False

	product_list = Product.objects.select_related("catalog").filter(catalog_id=current_category.id)
	if not count_products:
		count_products = product_list.count()
	wb = openpyxl.Workbook()
	ws = wb.active  #['Прайс']
	# w = xlwt.Workbook(style_compression=2)
	# ws = w.add_sheet(u'Прайс')
	HEAD_COLOR = 67

	# class Cell(object):
	# 	"""__init__() functions as the class constructor"""
	# 	def __init__(self, data=None, color=int()):
	# 		self.data = data
	# 		self.color = color

	# 	def getData(self):
	# 		return self.data

	# 	def getColor(self):
	# 		return self.color

	i = 1
	# style_date = xlwt.easyxf("pattern: pattern solid; alignment: wrap True, vertical top, horizontal left; border: left thin, right thin, top thin, bottom thin;", num_format_str='DD.MM.YYYY')

	ws.cell(row=i, column=1).value = "ID товара"
	ws.cell(row=i, column=2).value = "Приоритет"
	ws.cell(row=i, column=3).value = "Наименование товара"
	ws.cell(row=i, column=4).value = "Размер 1"
	ws.cell(row=i, column=5).value = "Размер 2"
	ws.cell(row=i, column=6).value = "Размер 3"
	ws.cell(row=i, column=7).value = "Размер 4"
	ws.cell(row=i, column=8).value = "Размер 5"
	ws.cell(row=i, column=9).value = "Размер 6"
	ws.cell(row=i, column=10).value = "Размер 7"
	ws.cell(row=i, column=11).value = "Марка"
	ws.cell(row=i, column=12).value = "ГОСТ/ТУ"
	ws.cell(row=i, column=13).value = "Наличие"
	ws.cell(row=i, column=14).value = "Цена"
	ws.cell(row=i, column=15).value = "Ед.изм"
	ws.cell(row=i, column=16).value = "Фото товара 1"
	ws.cell(row=i, column=17).value = "Фото товара 2"
	ws.cell(row=i, column=18).value = "Описание товара"
	ws.cell(row=i, column=19).value = "Производитель"
	ws.cell(row=i, column=20).value = "Страна производитель"
	ws.cell(row=i, column=21).value = "Способ изготовления"
	ws.cell(row=i, column=22).value = "ID категории товара"
	ws.cell(row=i, column=23).value = "Сертификат"
	ws.cell(row=i, column=24).value = "Доп. товар 1"
	ws.cell(row=i, column=25).value = "Доп. товар 2"
	ws.cell(row=i, column=26).value = "Доп. товар 3"
	ws.cell(row=i, column=27).value = "Доп. товар 4"
	ws.cell(row=i, column=28).value = "Услуга 1"
	ws.cell(row=i, column=29).value = "Услуга 2"
	ws.cell(row=i, column=30).value = "Услуга 3"
	ws.cell(row=i, column=31).value = "Услуга 4"
    # ws.write(i, 0, Cell(u"ID товара", HEAD_COLOR).data, style_date)


	ed_izm = ""
	for item in product_list:
		i = i + 1
		info.result_percent += (1. / count_products * 100)

		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or (i-1) == count_products:
			if flag_all:
				info.result = ("%.2f" % info.result_percent) + " %"
			else:
				info.result = ("%.0f" % info.result_percent) + " %"
			# info.info += "<p>В процессе - Экспорт категории: %s (id = %d) i=%d процент: %.2f</p>" % (current_category.name, item.id, i, info.result_percent)
			info.save()

		description = ""
		if (item.description):
			description=item.description

		marka=""
		if (item.marka_id):
			marka=item.marka

		standart=""
		if (item.gost_id):
			standart=item.gost
		# style_date.pattern.pattern_fore_colour = Cell(item.catalog, 1).color

		ws.cell(row=i, column=1).value = item.id
		if item.order_number:
			ws.cell(row=i, column=2).value = item.order_number
		else:
			ws.cell(row=i, column=2).value = ''
		ws.cell(row=i, column=3).value = item.name
		if item.size_a:
			ws.cell(row=i, column=4).value = item.size_a
		else:
			ws.cell(row=i, column=4).value = ""
		if item.size_b:
			ws.cell(row=i, column=5).value = item.size_b
		else:
			ws.cell(row=i, column=5).value = ""
		if item.size_c:
			ws.cell(row=i, column=6).value = item.size_c
		else:
			ws.cell(row=i, column=6).value = ""
		if item.size_d:
			ws.cell(row=i, column=7).value = item.size_d
		else:
			ws.cell(row=i, column=7).value = ""
		if item.size_e:
			ws.cell(row=i, column=8).value = item.size_e
		else:
			ws.cell(row=i, column=8).value = ""
		if item.size_f:
			ws.cell(row=i, column=9).value = item.size_f
		else:
			ws.cell(row=i, column=9).value = ""
		if item.size_l:
			ws.cell(row=i, column=10).value = item.size_l
		else:
			ws.cell(row=i, column=10).value = ""
		if item.marka:
			ws.cell(row=i, column=11).value = item.marka.name
		else:
			ws.cell(row=i, column=11).value = ""
		if item.gost:
			ws.cell(row=i, column=12).value = item.gost.number
		else:
			ws.cell(row=i, column=12).value = ""
		ws.cell(row=i, column=13).value = item.available
		ws.cell(row=i, column=14).value = item.price
		ws.cell(row=i, column=15).value = item.ed_izm

		if item.image and str(item.image)[-6:] != "/load/":
			ws.cell(row=i, column=16).value = '{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image) 
		else:
			ws.cell(row=i, column=16).value = ""

		if item.image_dop and str(item.image_dop)[-6:] != "/load/":
			ws.cell(row=i, column=17).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_dop)
		else:
			ws.cell(row=i, column=17).value = ""
		
		ws.cell(row=i, column=18).value = description
		ws.cell(row=i, column=19).value = item.vendor
		ws.cell(row=i, column=20).value = item.vendor_country
		ws.cell(row=i, column=21).value = item.method_of_manufacture
		ws.cell(row=i, column=22).value = item.catalog.id

		# if item.sertificate:
		# 	ws.cell(row=i, column=23).value = str(item.sertificate.id)
		# else:
		# 	ws.cell(row=i, column=23).value = ""
		ws.cell(row=i, column=23).value = ""

		# products_dop = tableProductDop.objects.filter(product=item)
		# for sub_item in products_dop[:4]:
		#     ws.write(i, tmp_ind, Cell(str(sub_item.product_dop.id)+".0", 1).data, style_date)
		#     tmp_ind += 1
		tmp_ind = 24
		while tmp_ind <= 27:
			ws.cell(row=i, column=tmp_ind).value = ""
			tmp_ind += 1

		# usluga_product = ProductUsluga.objects.filter(product=item)
		# tmp_ind = 27
		# for sub_item in usluga_product[:4]:
		#     ws.write(i, tmp_ind, Cell(str(sub_item.usluga.id)+".0", 1).data, style_date)
		#     tmp_ind += 1
		while tmp_ind <= 31:
			ws.cell(row=i, column=tmp_ind).value = ""
			tmp_ind += 1

    #  создание файла
	d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
	id = str(current_category.slug)
	name = id + "_" + d + ".xlsx"
	if not flag_all:
		info.link = name
	wb.save("www/media/export/" + name_all_dir + name)

def export_data_redirect(info, url_site, current_category=None, count_products=None, name_all_dir=""):
	flag_all = False

	product_list = Redirect.objects.all()
	count_products = product_list.count()
	wb = openpyxl.Workbook()
	ws = wb.active  #['Прайс']
	# w = xlwt.Workbook(style_compression=2)
	# ws = w.add_sheet(u'Прайс')
	HEAD_COLOR = 67

	# class Cell(object):
	# 	"""__init__() functions as the class constructor"""
	# 	def __init__(self, data=None, color=int()):
	# 		self.data = data
	# 		self.color = color

	# 	def getData(self):
	# 		return self.data

	# 	def getColor(self):
	# 		return self.color

	i = 1
	# style_date = xlwt.easyxf("pattern: pattern solid; alignment: wrap True, vertical top, horizontal left; border: left thin, right thin, top thin, bottom thin;", num_format_str='DD.MM.YYYY')

	ws.cell(row=i, column=1).value = "ID"
	ws.cell(row=i, column=2).value = "SITE"
	ws.cell(row=i, column=3).value = "Откуда"
	ws.cell(row=i, column=4).value = "Куда"
    # ws.write(i, 0, Cell(u"ID товара", HEAD_COLOR).data, style_date)


	ed_izm = ""
	for item in product_list:
		i = i + 1
		info.result_percent += (1. / count_products * 100)

		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or (i-1) == count_products:
			if flag_all:
				info.result = ("%.2f" % info.result_percent) + " %"
			else:
				info.result = ("%.0f" % info.result_percent) + " %"
			# info.info += "<p>В процессе - Экспорт категории: %s (id = %d) i=%d процент: %.2f</p>" % (current_category.name, item.id, i, info.result_percent)
			info.save()

		ws.cell(row=i, column=1).value = item.id
		ws.cell(row=i, column=2).value = "1"
		ws.cell(row=i, column=3).value = item.old_path
		ws.cell(row=i, column=4).value = item.new_path

    #  создание файла
	d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
	name = "redirect_" + d + ".xlsx"
	if not flag_all:
		info.link = name
	wb.save("www/media/export/" + name_all_dir + name)

def export_data(info, scheme):
    url_site = ''
    import time
    if info.redirect:
        export_data_redirect(info, url_site)
    elif info.name and info.name != DEFAULT_OPTION:
        export_data_category(info, url_site)
    else:
        export_all_category(info, url_site)
    info.state_id = IND_STATE_READY
    info.save()

    url_site = scheme + "://" + ProjectSettings.objects.first().site_name
    if info.email:
        link = url_site + info.get_full_link()
        link_admin = url_site + "/admin_m/export/"
        send_mail_result(info.email, info.user, info.name, link, link_admin, u"Экспорт", info.state.name)